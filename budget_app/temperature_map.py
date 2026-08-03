from __future__ import annotations

import os
import secrets
import shutil
import subprocess
import tempfile
import zipfile
from calendar import monthrange
from copy import copy
from dataclasses import dataclass
from datetime import date, datetime
from importlib.util import find_spec
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment
from .paths import data_path


OUTPUT_DIR = data_path(
    "saida/mapas_temperatura",
    "TEMPERATURE_MAP_OUTPUT_DIR",
    "saida/mapas_temperatura",
)
INACTIVE_MARK = "****"

MONTH_OPTIONS = [
    (1, "Janeiro"),
    (2, "Fevereiro"),
    (3, "Março"),
    (4, "Abril"),
    (5, "Maio"),
    (6, "Junho"),
    (7, "Julho"),
    (8, "Agosto"),
    (9, "Setembro"),
    (10, "Outubro"),
    (11, "Novembro"),
    (12, "Dezembro"),
]
MONTH_NAMES = {month: name for month, name in MONTH_OPTIONS}
MAP_TYPE_OPTIONS = [
    ("geral", "Mapa Temperatura - Geral"),
    ("geladeira", "Mapa Temperatura - Geladeira"),
    ("controlados", "Mapa Temperatura - Controlados"),
]
MAP_TYPES = {
    "geral": {
        "label": "Mapa Temperatura - Geral",
        "slug": "geral",
        "template_path": Path("modelos/mapa_temperatura/modelo_mapa_temperatura_medicamentos.xlsx"),
        "month_cell": "M7",
        "year_cell": "R7",
        "branch_cell": "U7",
        "first_day_row": 10,
        "last_day_row": 40,
        "first_variant_column": 2,
        "last_variant_column": 21,
    },
    "geladeira": {
        "label": "Mapa Temperatura - Geladeira",
        "slug": "geladeira",
        "template_path": Path("modelos/mapa_temperatura/modelo_mapa_temperatura_geladeira.xlsx"),
        "month_cell": "M7",
        "year_cell": "R7",
        "branch_cell": "U7",
        "first_day_row": 10,
        "last_day_row": 40,
        "first_variant_column": 2,
        "last_variant_column": 21,
    },
    "controlados": {
        "label": "Mapa Temperatura - Controlados",
        "slug": "controlados",
        "template_path": Path("modelos/mapa_temperatura/modelo_mapa_temperatura_controlados.xlsx"),
        "month_cell": "M7",
        "year_cell": "R7",
        "branch_cell": "U7",
        "first_day_row": 10,
        "last_day_row": 40,
        "first_variant_column": 2,
        "last_variant_column": 21,
    },
}


@dataclass(frozen=True)
class TemperatureMapInput:
    mes: int
    ano: int
    filial: str
    dias_extras: frozenset[int] = frozenset()

    @property
    def mes_nome(self) -> str:
        return MONTH_NAMES[self.mes]


def safe_filename_part(value: str) -> str:
    allowed = []
    for char in value.lower():
        if char.isalnum():
            allowed.append(char)
        elif char in {" ", "-", "_"}:
            allowed.append("_")
    return "".join(allowed).strip("_") or "mapa"


def parse_extra_days(values: list[str] | None) -> frozenset[int]:
    days = set()
    for value in values or []:
        value = value.strip()
        if not value:
            continue
        try:
            day = int(value)
        except ValueError as exc:
            raise ValueError("Campo Calendário: selecione apenas dias válidos.") from exc
        if day < 1 or day > 31:
            raise ValueError("Campo Calendário: selecione apenas dias entre 1 e 31.")
        days.add(day)
    return frozenset(days)


def parse_temperature_map_input(
    mes: str,
    ano: str,
    filial: str,
    dias_extras: list[str] | None = None,
) -> TemperatureMapInput:
    try:
        mes_int = int(mes)
    except ValueError as exc:
        raise ValueError("Campo Mês: selecione um mês válido.") from exc
    if mes_int not in MONTH_NAMES:
        raise ValueError("Campo Mês: selecione um mês válido.")

    try:
        ano_int = int(ano)
    except ValueError as exc:
        raise ValueError("Campo Ano: informe um ano válido.") from exc
    if ano_int < 2000 or ano_int > 2100:
        raise ValueError("Campo Ano: informe um ano entre 2000 e 2100.")

    filial_digits = "".join(char for char in filial if char.isdigit())
    if len(filial_digits) != 3:
        raise ValueError("Campo Filial: informe exatamente 3 dígitos.")

    days_in_month = monthrange(ano_int, mes_int)[1]
    valid_extra_days = [
        day
        for day in parse_extra_days(dias_extras)
        if day <= days_in_month
    ]

    return TemperatureMapInput(
        mes=mes_int,
        ano=ano_int,
        filial=filial_digits,
        dias_extras=frozenset(valid_extra_days),
    )


def parse_map_types(values: list[str] | None) -> list[str]:
    selected = [value.strip() for value in values or [] if value.strip()]
    if not selected:
        raise ValueError("Selecione pelo menos um mapa.")
    invalid = [value for value in selected if value not in MAP_TYPES]
    if invalid:
        raise ValueError("Campo Modelo: selecione apenas mapas válidos.")
    return list(dict.fromkeys(selected))


def ensure_temperature_template(map_type: str) -> Path:
    template_path = MAP_TYPES[map_type]["template_path"]
    if not template_path.exists():
        label = MAP_TYPES[map_type]["label"]
        raise FileNotFoundError(f"Modelo XLSX não encontrado para {label}: {template_path}")
    return template_path


def log_temperature_map(message: str) -> None:
    print(f"[mapa-temperatura] {message}", flush=True)


def xlsx_package_diagnostics(path: Path) -> dict[str, object]:
    info: dict[str, object] = {
        "path": str(path),
        "exists": path.exists(),
    }
    if not path.exists():
        return info

    info["size"] = path.stat().st_size
    try:
        with zipfile.ZipFile(path) as archive:
            names = archive.namelist()
            media = sorted(name for name in names if name.startswith("xl/media/"))
            drawings = sorted(name for name in names if name.startswith("xl/drawings/"))
            rels = sorted(
                name
                for name in names
                if name.startswith("xl/worksheets/_rels/") or name.startswith("xl/drawings/_rels/")
            )
            info["media_count"] = len(media)
            info["media"] = media
            info["drawing_count"] = len(drawings)
            info["drawing_files"] = drawings
            info["relationship_files"] = rels
    except Exception as exc:
        info["zip_error"] = str(exc)
    return info


def log_xlsx_diagnostics(label: str, path: Path) -> dict[str, object]:
    info = xlsx_package_diagnostics(path)
    media = ", ".join(info.get("media", []) or ["sem mídia"])
    drawings = ", ".join(info.get("drawing_files", []) or ["sem drawings"])
    log_temperature_map(
        f"{label}: exists={info.get('exists')} size={info.get('size')} "
        f"media_count={info.get('media_count')} media={media} "
        f"drawing_count={info.get('drawing_count')} drawings={drawings}"
    )
    if info.get("zip_error"):
        log_temperature_map(f"{label}: erro ao inspecionar xlsx: {info['zip_error']}")
    return info


def count_openpyxl_images(workbook) -> int:
    total = 0
    for worksheet in workbook.worksheets:
        total += len(getattr(worksheet, "_images", []) or [])
    return total


def summarize_dimensions(worksheet, config: dict[str, object]) -> str:
    rows = []
    for row in range(1, int(config["last_day_row"]) + 1):
        height = worksheet.row_dimensions[row].height
        if height is not None:
            rows.append(f"{row}:{height:g}")
    cols = []
    for col, dimension in sorted(worksheet.column_dimensions.items()):
        if dimension.width is not None:
            cols.append(f"{col}:{dimension.width:g}")
    return (
        f"default_row_height={worksheet.sheet_format.defaultRowHeight} "
        f"rows={';'.join(rows) or 'sem_alturas_customizadas'} "
        f"cols={';'.join(cols) or 'sem_larguras_customizadas'}"
    )


def log_print_setup(worksheet, label: str) -> None:
    setup = worksheet.page_setup
    margins = worksheet.page_margins
    log_temperature_map(
        f"{label}: orientation={setup.orientation} paperSize={setup.paperSize} "
        f"scale={setup.scale} fitToWidth={setup.fitToWidth} fitToHeight={setup.fitToHeight} "
        f"margins=(left={margins.left}, right={margins.right}, top={margins.top}, bottom={margins.bottom})"
    )


def lock_temperature_layout(worksheet, config: dict[str, object]) -> None:
    default_height = worksheet.sheet_format.defaultRowHeight or 15
    last_row = int(config["last_day_row"])
    for row in range(1, last_row + 1):
        dimension = worksheet.row_dimensions[row]
        dimension.height = dimension.height or default_height
    log_temperature_map(f"layout fixado: {summarize_dimensions(worksheet, config)}")


def inactive_temperature_days(data: TemperatureMapInput) -> set[int]:
    days_in_month = monthrange(data.ano, data.mes)[1]
    inactive_days = set(data.dias_extras)
    inactive_days.update(day for day in range(days_in_month + 1, 32))
    for day in range(1, days_in_month + 1):
        if date(data.ano, data.mes, day).weekday() == 6:
            inactive_days.add(day)
    return inactive_days


def center_cell(cell) -> None:
    current = copy(cell.alignment)
    cell.alignment = Alignment(
        horizontal="center",
        vertical="center",
        text_rotation=current.textRotation,
        wrap_text=current.wrap_text,
        shrink_to_fit=current.shrink_to_fit,
        indent=current.indent,
    )


def fill_inactive_temperature_rows(worksheet, data: TemperatureMapInput, map_type: str) -> None:
    config = MAP_TYPES[map_type]
    inactive_days = inactive_temperature_days(data)
    days_in_month = monthrange(data.ano, data.mes)[1]
    for day in range(1, 32):
        row = config["first_day_row"] + day - 1
        worksheet.row_dimensions[row].hidden = day > days_in_month
    for day in inactive_days:
        row = config["first_day_row"] + day - 1
        if row > config["last_day_row"]:
            continue
        for column in range(config["first_variant_column"], config["last_variant_column"] + 1):
            cell = worksheet.cell(row=row, column=column)
            cell.value = INACTIVE_MARK
            center_cell(cell)


def render_temperature_map(data: TemperatureMapInput, map_type: str = "geral") -> Path:
    config = MAP_TYPES[map_type]
    template_path = ensure_temperature_template(map_type)
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    pillow_available = find_spec("PIL") is not None
    log_temperature_map(
        f"iniciando geração: tipo={map_type} mês={data.mes} ano={data.ano} "
        f"filial={data.filial} pillow={pillow_available}"
    )
    template_info = log_xlsx_diagnostics("modelo antes do openpyxl", template_path)
    workbook = load_workbook(template_path)
    log_temperature_map(f"imagens carregadas pelo openpyxl={count_openpyxl_images(workbook)}")
    worksheet = workbook.active
    log_temperature_map(f"layout modelo: {summarize_dimensions(worksheet, config)}")
    log_print_setup(worksheet, "impressão modelo")
    worksheet[config["month_cell"]] = data.mes_nome
    worksheet[config["year_cell"]] = data.ano
    worksheet[config["branch_cell"]] = int(data.filial)
    worksheet[config["branch_cell"]].number_format = "000"
    center_cell(worksheet[config["branch_cell"]])
    fill_inactive_temperature_rows(worksheet, data, map_type)
    lock_temperature_layout(worksheet, config)
    log_print_setup(worksheet, "impressão antes de salvar")

    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    nonce = secrets.token_hex(3)
    filename = (
        f"mapa_temperatura_{config['slug']}_filial_{data.filial}_"
        f"{safe_filename_part(data.mes_nome)}_{data.ano}_{stamp}_{nonce}.xlsx"
    )
    output_path = OUTPUT_DIR / filename
    workbook.save(output_path)
    output_info = log_xlsx_diagnostics("arquivo gerado depois do openpyxl", output_path)
    if (template_info.get("media_count") or 0) > 0 and (output_info.get("media_count") or 0) == 0:
        log_temperature_map(
            "alerta: o modelo possui imagem, mas o arquivo gerado ficou sem mídia. "
            "Verifique se Pillow foi instalado no ambiente do Render."
        )
    return output_path


def powershell_quote(value: str) -> str:
    return "'" + value.replace("'", "''") + "'"


def convert_xlsx_to_pdf_with_excel(xlsx_path: Path) -> tuple[Path | None, str]:
    log_temperature_map(f"pdf excel: início arquivo={xlsx_path.name}")
    if os.name != "nt":
        log_temperature_map("pdf excel: ignorado porque ambiente não é Windows")
        return None, "Microsoft Excel disponível apenas no Windows."

    powershell = shutil.which("powershell") or shutil.which("pwsh")
    if not powershell:
        log_temperature_map("pdf excel: PowerShell não encontrado")
        return None, "PowerShell não encontrado para acionar o Microsoft Excel."

    source_path = xlsx_path.resolve()
    pdf_path = xlsx_path.with_suffix(".pdf")
    temp_path = pdf_path.with_name(f"{pdf_path.stem}.tmp.pdf")
    temp_path.unlink(missing_ok=True)

    script = f"""
$ErrorActionPreference = 'Stop'
$xlsxPath = {powershell_quote(str(source_path))}
$pdfPath = {powershell_quote(str(temp_path.resolve()))}
if (-not (Test-Path -LiteralPath $xlsxPath)) {{
  throw "XLSX não encontrado: $xlsxPath"
}}
if (Test-Path -LiteralPath $pdfPath) {{
  Remove-Item -LiteralPath $pdfPath -Force
}}
$excel = $null
$workbook = $null
try {{
  $excel = New-Object -ComObject Excel.Application
  $excel.Visible = $false
  $excel.DisplayAlerts = $false
  $workbook = $excel.Workbooks.Open($xlsxPath, 3, $true)
  $workbook.ExportAsFixedFormat(0, $pdfPath)
}} finally {{
  if ($workbook -ne $null) {{
    $workbook.Close($false) | Out-Null
  }}
  if ($excel -ne $null) {{
    $excel.Quit() | Out-Null
  }}
  [System.GC]::Collect()
  [System.GC]::WaitForPendingFinalizers()
}}
"""
    result = subprocess.run(
        [
            powershell,
            "-NoProfile",
            "-NonInteractive",
            "-ExecutionPolicy",
            "Bypass",
            "-Command",
            script,
        ],
        check=False,
        capture_output=True,
        text=True,
        timeout=120,
    )
    if result.returncode == 0 and temp_path.exists():
        temp_path.replace(pdf_path)
        log_temperature_map(f"pdf excel: sucesso arquivo={pdf_path.name}")
        return pdf_path, "PDF gerado com sucesso pelo Microsoft Excel."

    temp_path.unlink(missing_ok=True)
    detail = (result.stderr or result.stdout or "erro desconhecido").strip()
    log_temperature_map(f"pdf excel: falha returncode={result.returncode} detalhe={compact_error_detail(detail, 500)}")
    return None, f"Microsoft Excel não gerou o PDF: {detail}"


def find_soffice() -> str | None:
    found = shutil.which("soffice") or shutil.which("libreoffice")
    if found:
        return found
    for candidate in (
        Path("C:/Program Files/LibreOffice/program/soffice.exe"),
        Path("C:/Program Files (x86)/LibreOffice/program/soffice.exe"),
        Path.home() / "AppData/Local/Programs/LibreOffice/program/soffice.exe",
    ):
        if candidate.exists():
            return str(candidate)
    return None


def compact_error_detail(value: str, limit: int = 1200) -> str:
    text = " ".join((value or "erro desconhecido").split())
    if len(text) <= limit:
        return text
    return text[: limit - 3] + "..."


def soffice_version(soffice: str) -> str:
    try:
        result = subprocess.run(
            [soffice, "--version"],
            check=False,
            capture_output=True,
            text=True,
            timeout=15,
        )
    except Exception as exc:
        return f"erro ao consultar versão: {exc}"
    return compact_error_detail(result.stdout or result.stderr, 300)


def convert_xlsx_to_pdf_with_libreoffice(xlsx_path: Path) -> tuple[Path | None, str]:
    log_temperature_map(f"pdf libreoffice: início arquivo={xlsx_path.name}")
    soffice = find_soffice()
    if not soffice:
        log_temperature_map("pdf libreoffice: soffice não encontrado")
        return None, "LibreOffice/soffice não encontrado."
    log_temperature_map(f"pdf libreoffice: soffice={soffice} versão={soffice_version(soffice)}")

    source_path = xlsx_path.resolve()
    output_dir = xlsx_path.parent.resolve()
    pdf_path = xlsx_path.with_suffix(".pdf")
    pdf_path.unlink(missing_ok=True)

    try:
        with tempfile.TemporaryDirectory(prefix="lo-profile-") as profile:
            profile_dir = Path(profile).resolve()
            env = {
                **os.environ,
                "HOME": os.environ.get("HOME", "/tmp"),
                "SAL_USE_VCLPLUGIN": os.environ.get("SAL_USE_VCLPLUGIN", "svp"),
                "SAL_DISABLE_OPENCL": os.environ.get("SAL_DISABLE_OPENCL", "1"),
                "SAL_DISABLE_GL": os.environ.get("SAL_DISABLE_GL", "1"),
                "JAVA_TOOL_OPTIONS": os.environ.get("JAVA_TOOL_OPTIONS", "-Xmx96m"),
            }
            command = [
                soffice,
                "--headless",
                "--nologo",
                "--nodefault",
                "--nolockcheck",
                "--norestore",
                "--nofirststartwizard",
                f"-env:UserInstallation={profile_dir.as_uri()}",
                "--convert-to",
                "pdf",
                "--outdir",
                str(output_dir),
                str(source_path),
            ]
            log_temperature_map(
                "pdf libreoffice: executando conversão "
                f"timeout=75s output_dir={output_dir} profile={profile_dir}"
            )
            result = subprocess.run(
                command,
                check=False,
                capture_output=True,
                text=True,
                timeout=75,
                env=env,
            )
            log_temperature_map(f"pdf libreoffice: subprocess finalizado returncode={result.returncode}")
    except subprocess.TimeoutExpired as exc:
        log_temperature_map("pdf libreoffice: timeout após 75s")
        return None, f"LibreOffice não gerou o PDF: tempo limite excedido. {exc}"
    except Exception as exc:
        log_temperature_map(f"pdf libreoffice: exceção={compact_error_detail(str(exc), 500)}")
        return None, f"LibreOffice não gerou o PDF: {exc}"

    if result.returncode == 0 and pdf_path.exists():
        log_temperature_map(f"pdf libreoffice: sucesso arquivo={pdf_path.name} stdout={compact_error_detail(result.stdout, 300)}")
        return pdf_path, "PDF gerado com sucesso pelo LibreOffice."
    detail = compact_error_detail(result.stderr or result.stdout)
    log_temperature_map(f"pdf libreoffice: falha returncode={result.returncode} detalhe={compact_error_detail(detail, 500)}")
    return None, f"LibreOffice não gerou o PDF: {detail}"


def convert_xlsx_to_pdf_with_docker(xlsx_path: Path) -> tuple[Path | None, str]:
    log_temperature_map(f"pdf docker: início arquivo={xlsx_path.name}")
    docker = shutil.which("docker")
    if not docker:
        log_temperature_map("pdf docker: docker não encontrado")
        return None, "Docker não encontrado."

    image = os.environ.get("LIBREOFFICE_DOCKER_IMAGE", "orcamento-clamed-local")
    output_dir = xlsx_path.parent.resolve()
    pdf_path = xlsx_path.with_suffix(".pdf")
    pdf_path.unlink(missing_ok=True)
    container_file = f"/work/{xlsx_path.name}"

    try:
        result = subprocess.run(
            [
                docker,
                "run",
                "--rm",
                "-v",
                f"{output_dir}:/work",
                "-w",
                "/work",
                "--entrypoint",
                "soffice",
                image,
                "--headless",
                "--nologo",
                "--nodefault",
                "--nolockcheck",
                "--norestore",
                "--nofirststartwizard",
                "--convert-to",
                "pdf",
                "--outdir",
                "/work",
                container_file,
            ],
            check=False,
            capture_output=True,
            text=True,
            timeout=75,
        )
    except subprocess.TimeoutExpired as exc:
        return None, f"Docker/LibreOffice não gerou o PDF: tempo limite excedido. {exc}"
    except Exception as exc:
        log_temperature_map(f"pdf docker: exceção={compact_error_detail(str(exc), 500)}")
        return None, f"Docker/LibreOffice não gerou o PDF: {exc}"

    if result.returncode == 0 and pdf_path.exists():
        log_temperature_map(f"pdf docker: sucesso arquivo={pdf_path.name}")
        return pdf_path, "PDF gerado com sucesso pelo LibreOffice no Docker."
    detail = compact_error_detail(result.stderr or result.stdout)
    log_temperature_map(f"pdf docker: falha returncode={result.returncode} detalhe={compact_error_detail(detail, 500)}")
    return None, f"Docker/LibreOffice não gerou o PDF: {detail}"


def convert_temperature_map_to_pdf(xlsx_path: Path) -> tuple[Path | None, str]:
    errors = []
    for converter in (
        convert_xlsx_to_pdf_with_libreoffice,
        convert_xlsx_to_pdf_with_docker,
        convert_xlsx_to_pdf_with_excel,
    ):
        log_temperature_map(f"pdf fluxo: tentando conversor={converter.__name__} arquivo={xlsx_path.name}")
        pdf_path, status = converter(xlsx_path)
        if pdf_path is not None:
            log_temperature_map(f"pdf fluxo: sucesso conversor={converter.__name__} status={status}")
            return pdf_path, status
        log_temperature_map(f"pdf fluxo: falha conversor={converter.__name__} status={compact_error_detail(status, 700)}")
        errors.append(status)
    detail = " | ".join(errors)
    if os.environ.get("RENDER") and "LibreOffice/soffice não encontrado" in detail:
        return None, (
            "PDF não gerado no Render porque o serviço atual não está usando a imagem Docker "
            "com LibreOffice. O XLSX foi criado normalmente. No painel do Render, crie/recrie "
            "o serviço como Docker Web Service ou aplique o Blueprint render.yaml. "
            f"Detalhes: {compact_error_detail(detail, 1400)}"
        )
    return None, (
        "PDF não gerado no servidor. O XLSX foi criado normalmente. "
        "Detalhes: "
        f"{compact_error_detail(detail, 1800)}"
    )


def output_file(filename: str) -> Path | None:
    if not filename or Path(filename).name != filename:
        return None
    path = OUTPUT_DIR / filename
    try:
        path.resolve().relative_to(OUTPUT_DIR.resolve())
    except ValueError:
        return None
    return path if path.exists() else None


def list_temperature_maps() -> list[Path]:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    return sorted(OUTPUT_DIR.glob("*.xlsx"), key=lambda path: path.stat().st_mtime, reverse=True)
